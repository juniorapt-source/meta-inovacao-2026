# -*- coding: utf-8 -*-
"""Gera os arquivos data/*.js a partir do plano v5 + conteúdos do e-mail/caminho crítico.
Formato: window.DB.<chave> = <JSON>;  (JSON puro após o '=' — validável por script)"""
import json, re, io, os

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
from openpyxl import load_workbook
_ws = load_workbook('/home/claude/work/Plano_de_ac_a_o_v5.xlsx', read_only=True)['Plano de Ação']
SRC = []
for _i, _row in enumerate(_ws.iter_rows(values_only=True)):
    if _i == 0 or not _row[0]: continue
    SRC.append({"id": _row[0], "frente": _row[1], "sub": _row[2], "atividade": _row[3],
                "resp": _row[4], "prazo": _row[5], "status": _row[6], "dep": _row[7], "cc": _row[8]})

# ---------- soluções por ação: como executar · como monitorar · ferramenta ----------
S = {
 "CMT-01": ("Concluída. Registrar em ata a composição e os três papéis do Comitê (engajamento, curadoria, devolutiva) para citação nas comunicações.",
            "Ata publicada e referenciada no boletim CG-01.", "Ata + repositório do projeto"),
 "CMT-02": ("Sandra cria série recorrente no calendário (semanal ou quinzenal, 50 min) com pauta padrão fixa: nós da semana, pendências de formulário, decisões.",
            "Série criada no calendário e taxa de realização das reuniões registrada neste painel (agenda).", "Outlook/Teams + painel"),
 "CMT-03": ("Usar o deck já validado; apresentar plano, papéis e este painel na reunião de equipe de 10/08. Encerrar com o trilho dos 7 nós.",
            "Apresentação realizada; link do painel circulado ao time.", "Deck + painel (link Vercel)"),
 "CMT-04": ("Definir no rito de 07/08: pauta padrão, quem registra decisões, prazo de resposta do Comitê (ex.: 2 d.u.) e canal único de entrada de demandas.",
            "Documento de rito de 1 página versionado no repo (docs/).", "docs/rito_comite.md no repo"),
 "INS-01": ("Concluída. Lote 1 definido com as unidades-alvo (UASJUR, UGE, UGS/Soluções, URC).",
            "Lista do Lote 1 refletida nas ações INS-02/06/07.", "Painel · plano"),
 "INS-02": ("Sandra agenda os debates restantes (Op. Fomento e Op. Investimento) até 11/08; convocar Agnaldo e, se preciso, Giovanni para a fronteira FAMPE/Finep.",
            "Debates realizados até 11/08 = Nó 3 no verde; registrar resultado por instrumento no painel.", "Agenda + painel (caminho crítico)"),
 "INS-03": ("Concluída. Lote 2 fechado (inclui itens dependentes da UDT e definições ALI IG).",
            "Escopo do Lote 2 listado no painel.", "Painel · plano"),
 "INS-04": ("Anny confirma a oficina UDT (convite saiu até 07/08). Antecipar com o time o que não depende da UDT (ex.: ALI IG) nas mesmas agendas.",
            "Data da oficina cravada até 14/08; sem data → gatilho do Nó 6 (gerência aciona).", "Convite Anny + gatilho Nó 6"),
 "INS-06": ("Validação final do Lote 1 no Comitê em 12/08; se Fomento/Investimento não fecharem, aplicar fallback: Lote 1A (3 prontos) sobe dia 14, restante 21/08.",
            "Decisão registrada em ata; status muda para Concluído no painel em 12/08.", "Rito do Comitê + painel"),
 "INS-07": ("Submeter ofícios/fichas às unidades em 14/08 (compromisso público do deck — não desliza). Distribuir a carga do dia com Sandra e Pova.",
            "Protocolos de submissão anexados; Nó 4 verde no painel em 14/08.", "Ofício/e-mail institucional + painel"),
 "INS-08": ("Validar Lote 2 no Comitê em 18/09, com insumos da oficina UDT (Nó 6).",
            "Ata de validação; cadeia 6 em dia no painel.", "Rito do Comitê"),
 "INS-09": ("Submeter Lote 2 em 25/09 às unidades responsáveis, mesmo rito do Lote 1.",
            "Protocolos anexados; cadeia 6 concluída.", "Ofício/e-mail institucional"),
 "INS-10": ("Consolidar devolutivas por instrumento numa tabela única (aceito · aceito com ajustes · não aceito) e incorporar às fichas.",
            "Tabela de devolutivas preenchida; % de fichas ajustadas visível no boletim.", "Planilha de devolutivas + boletim CG-01"),
 "INS-11": ("Comunicado às UFs sobre o Lote 1 aprovado: o que muda, a partir de quando, onde registrar.",
            "Comunicado enviado; data registrada no painel.", "Marketing Cloud / canal institucional"),
 "INS-12": ("Mesmo modelo do INS-11 para o Lote 2.",
            "Comunicado enviado; data registrada no painel.", "Marketing Cloud / canal institucional"),
 "CAN-01": ("Concluída. Proposta enviada à URC (Enio e Milva) com estrutura de 2 sessões × 3h por canal.",
            "E-mail enviado em 30/07.", "E-mail institucional"),
 "CAN-02": ("Aguardar devolutiva da URC; silêncio até 12h de 05/08 já venceu → manter ligação gerência-a-gerência ativa. Qualquer contraproposta serve se o Ciclo 1 começar até ~18/08.",
            "Nó 1: grade devolvida = verde. Cada dia sem resposta consome a folga de 26/08–04/09.", "Gerência + gatilho Nó 1"),
 "CAN-03": ("Gerência apresenta plano, objetivos das oficinas e cronograma na reunião de equipe de 10/08 (mesma agenda do CMT-03).",
            "Apresentação realizada; presença dos coordenadores registrada.", "Reunião de equipe + painel"),
 "CAN-04": ("Com a grade validada, Sandra dispara calendário + convites de agenda + materiais em até 24h (meta 11/08, limite 13/08).",
            "Convites aceitos por canal ≥ mínimo definido; página agenda do painel preenchida.", "Outlook + painel (agenda)"),
 "CAN-05": ("Cada núcleo indica representante até 07/08; sem indicação, vale o fallback aprovado: o gestor da iniciativa é o representante.",
            "5/5 núcleos com nome na página participantes (pendências ficam vermelhas).", "Rito do Comitê + painel (participantes)"),
 "CAN-06": ("Sandra reserva salas para todos os encontros; sala indisponível → canal migra para online (DXP, 2h, é candidato natural — pré-acordar com a URC).",
            "Encontros com local definido na agenda do painel; pendências visíveis.", "Reserva de salas + painel (agenda)"),
 "CAN-07": ("Montar repositório por canal (materiais da URC + fichas das iniciativas + Diretrizes v1.2) para alimentar o agente de IA.",
            "Repositório com estrutura mínima por canal até 11/08.", "SharePoint/Drive + link no painel"),
 "CAN-08": ("JR. e Pova desenvolvem o agente de apoio ao preenchimento do formulário URC usando o repositório do CAN-07.",
            "Agente responde às perguntas do formulário com dados de 1 iniciativa de teste.", "Agente IA (Pova)"),
 "CAN-09": ("Piloto com 1–2 iniciativas em 14/08. O agente está FORA do caminho crítico: o plano B (formulário estático) cobre o Ciclo 1.",
            "Piloto rodado e ajustes listados; decisão go/no-go para uso no Ciclo 1.", "Agente IA + formulário plano B"),
 "CAN-10": ("Cada projeto entrega o formulário (ou justificativa de não adoção) em até 5 d.u. após cada oficina. C1 até 01/09, C2 até 30/09.",
            "Matriz de demandas do painel: célula muda de 'oficina' para 'formulário' dentro do prazo; checkpoint 19/08 (5 de 8 canais rodados).",
            "Formulário URC + matriz do painel"),
 "CAN-11": ("Sandra monitora confirmações de presença, alimenta o repositório e cobra pendências semanalmente (rotina fixa, ex.: sexta 10h).",
            "Página agenda com % de confirmação por encontro; pendências nominais no boletim.", "Painel (agenda) + boletim CG-01"),
 "CAN-12": ("Comitê define critérios de priorização até 21/08; fallback barato: primeira pauta do Comitê após o início do ciclo.",
            "Documento de critérios (1 página) aprovado em ata; sem ele, Nó 7 fica âmbar.", "Rito do Comitê + docs/ no repo"),
 "CAN-13": ("Aplicar os critérios às demandas do Ciclo 1 e encaminhar lote priorizado à URC em 11/09 (Nó 7).",
            "Lote enviado com nº de demandas por canal; células da matriz mudam para 'priorizado'.", "Comitê + matriz do painel"),
 "CAN-14": ("Mesmo rito para o Ciclo 2, encaminhando até 09/10.",
            "Lote 2 enviado; matriz atualizada.", "Comitê + matriz do painel"),
 "CAN-15": ("Consolidar devolutivas por canal + diagnóstico de aderência (adotado · em adoção · não adotado com justificativa) — insumo direto do RD/Adequação.",
            "Relatório de aderência por canal publicado em 16/10.", "Painel (matriz) + relatório"),
 "CAN-16": ("Gerência comunica aos Estados as demandas priorizadas e o que muda para as UFs.",
            "Comunicado enviado em out; registrado no painel.", "Canal institucional NA→UF"),
 "CAN-17": ("Negociar o SLA na própria devolutiva do Nó 1: URC devolve cronograma em até 10 d.u. após cada lote priorizado.",
            "Cronograma da URC anexado; sem SLA, a linha fica âmbar permanente no painel.", "Acordo URC (SLA) + painel"),
 "CAN-18": ("Sandra acompanha entregas da URC × pendências da UI em rotina quinzenal, alimentando a matriz (célula 'encaminhado').",
            "Matriz com datas de entrega; desvios sobem para o Comitê.", "Painel (matriz) + rito do Comitê"),
 "PTF-01": ("Enviar em 14/08 o pedido das 4 oficinas à Soluções já completo: proposta de formato junto (PTF-02) e datas sugeridas 28/09–09/10 (fora do fim do Ciclo 2).",
            "Pedido protocolado em 14/08; resposta da Soluções com datas = SLA ok.", "Ofício + painel (caminho crítico)"),
 "PTF-02": ("Elaborar a proposta de formato (espelhando a estrutura URC: apresentação + oficina + devolutiva macro) para anexar ao PTF-01.",
            "Documento de formato anexado ao pedido de 14/08.", "Documento (docs/) + ofício"),
 "PTF-03": ("Apresentar o plano das oficinas de Soluções na reunião de equipe de setembro.",
            "Apresentação realizada; registrada no painel.", "Reunião de equipe"),
 "PTF-04": ("Comitê indica representantes (aproveitar a mesma lógica e nomes do CAN-05 quando fizer sentido).",
            "Indicações completas na página participantes.", "Rito do Comitê + painel"),
 "PTF-05": ("Sandra reserva locais assim que a Soluções confirmar as datas.",
            "Locais definidos na agenda do painel.", "Reserva de salas + painel"),
 "PTF-06": ("Formulários de criação/ajuste de soluções entregues em até 5 d.u. após cada oficina (mesma régua do CAN-10).",
            "Controle por núcleo no painel; pendências no boletim.", "Formulário Soluções + painel"),
 "PTF-07": ("Comitê prioriza e encaminha as demandas de soluções em outubro, com os critérios do CAN-12 adaptados.",
            "Lote encaminhado; registrado no painel.", "Comitê"),
 "PTF-08": ("Gerência comunica aos Estados as demandas de portfólio.",
            "Comunicado enviado out–nov.", "Canal institucional NA→UF"),
 "PTF-09": ("Receber da Soluções o cronograma de entrega (SLA espelho do CAN-17).",
            "Cronograma anexado; sem ele, âmbar no painel.", "Acordo Soluções (SLA)"),
 "PTF-10": ("Sandra monitora entregas de Soluções × pendências da UI, rotina quinzenal.",
            "Status por demanda no painel; desvios ao Comitê.", "Painel + rito do Comitê"),
 "CG-01": ("Boletim mensal de 1 página: nós, entregas do mês, pendências nominais, próximos marcos. 1ª edição 31/08. Fonte: este painel (imprimir/exportar).",
            "Boletim enviado todo fim de mês; taxa de abertura se via Marketing Cloud.", "Painel + Marketing Cloud"),
 "CG-02": ("Sandra atualiza o status do projeto em toda reunião de equipe usando o dashboard do painel projetado.",
            "Item fixo de pauta cumprido; registrado em ata.", "Painel (dashboard)"),
 "CG-03": ("Este projeto: publicar o painel no Vercel via repositório Git (deploy automático a cada commit) até 14/08.",
            "URL pública ativa; versão e data de atualização no rodapé de todas as páginas.", "Vercel + Git"),
 "CG-04": ("Retrospectiva ao final do Ciclo 2: o que protegeu os nós, o que consumiu folga, o que muda para 2027.",
            "Documento de lições aprendidas versionado no repo (docs/).", "docs/retrospectiva.md"),
}

MESES = {"jan":1,"fev":2,"mar":3,"abr":4,"mai":5,"jun":6,"jul":7,"ago":8,"set":9,"out":10,"nov":11,"dez":12}

def prazo_iso(p):
    if not p: return None
    p = str(p).strip()
    m = re.fullmatch(r"(\d{1,2})/(\d{1,2})", p)
    if m:
        d, mo = int(m.group(1)), int(m.group(2))
        return f"2026-{mo:02d}-{d:02d}"
    return None

def parse_no(cc):
    if not cc: return {"tipo": None}
    cc = str(cc)
    m = re.search(r"Nó (\d)", cc)
    if m: return {"tipo": "no", "no": int(m.group(1))}
    if "cadeia" in cc: return {"tipo": "cadeia", "no": int(re.search(r"\d", cc).group())}
    if "SLA" in cc: return {"tipo": "sla"}
    return {"tipo": None}

plano = []
for r in SRC:
    dep = [] if (r["dep"] in (None, "—", "")) else [d.strip() for d in str(r["dep"]).replace("·", "|").split("|")]
    sol = S[r["id"]]
    plano.append({
        "id": r["id"], "frente": r["frente"], "sub": r["sub"], "atividade": r["atividade"],
        "resp": r["resp"], "prazo": ("" if r["prazo"] in (None,) else str(r["prazo"]).strip()),
        "prazo_iso": prazo_iso(r["prazo"]), "status": r["status"] or "Não iniciado",
        "dep": dep, "cc": parse_no(r["cc"]),
        "como": sol[0], "monitor": sol[1], "ferramenta": sol[2],
    })
assert len(plano) == 47, len(plano)
ids = {a["id"] for a in plano}
for a in plano:
    for d in a["dep"]:
        assert d in ids, (a["id"], d)

nos = [
 {"no":1,"data":"04/08","data_iso":"2026-08-04","titulo":"Devolutiva da URC sobre a grade","acoes":["CAN-02"],"guardiao":"Gerência",
  "fallback":"Qualquer contraproposta de datas serve, desde que o Ciclo 1 comece até ~18/08 — depois disso a priorização de 11/09 desliza em cadeia.",
  "gatilho":"Silêncio até 12h de 05/08 → ligação gerência-a-gerência no mesmo dia."},
 {"no":2,"data":"05–07/08","data_iso":"2026-08-07","titulo":"Representantes indicados + locais reservados","acoes":["CAN-05","CAN-06"],"guardiao":"Comitê e Sandra",
  "fallback":"Aprovado no rito de 07/08: sem indicação do núcleo, o gestor da iniciativa é o representante; sala indisponível → canal migra para online (pré-acordar com a URC; DXP, de 2h, é candidato natural).",
  "gatilho":"Pendência na sexta 07/08 → aplicar fallback automaticamente."},
 {"no":3,"data":"12/08","data_iso":"2026-08-12","titulo":"Fechamento do Lote 1 (debates + validação do Comitê)","acoes":["INS-02","INS-06"],"guardiao":"JR. e gestores + Comitê",
  "fallback":"Submissão modular — os 3 instrumentos prontos sobem dia 14 como Lote 1A e o restante em 21/08.",
  "gatilho":"Debates de Fomento/Investimento não realizados até 11/08 → acionar fallback."},
 {"no":4,"data":"14/08","data_iso":"2026-08-14","titulo":"Submissão do Lote 1 às unidades","acoes":["INS-07"],"guardiao":"JR./Gerência",
  "fallback":"Compromisso público do deck (slides 10 e 17): não desliza. É a data que os fallbacks do Nó 3 protegem.",
  "gatilho":"14/08 concentra também PTF-01, PTF-02, CAN-09 e CG-03 — distribuir execução com Sandra e Pova."},
 {"no":5,"data":"11→13/08","data_iso":"2026-08-13","titulo":"Calendário + convites + formulário plano B antes do Ciclo 1","acoes":["CAN-04","CAN-10"],"guardiao":"Sandra + gestores",
  "fallback":"Formulário estático com opções estruturadas pronto em 11/08 — o Ciclo 1 abre coberto por ele. O agente de IA (CAN-08/09) fica fora do caminho crítico.",
  "gatilho":"Checkpoint 19/08 (5 dos 8 canais rodados): presença baixa ou formulários travados → coordenador do núcleo age na mesma semana."},
 {"no":6,"data":"≤04/09","data_iso":"2026-09-04","titulo":"Oficina com a UDT (encadeia Lote 2)","acoes":["INS-04"],"guardiao":"Sandra/Anny + JR.",
  "fallback":"O que não depende da UDT (ex.: definições do ALI IG) é antecipado nas agendas do Lote 2. Convite precisou sair até 07/08.",
  "gatilho":"14/08 sem data da oficina → gerência aciona. Hoje o nó mais perigoso: dependência sem data consome folga invisivelmente."},
 {"no":7,"data":"11/09","data_iso":"2026-09-11","titulo":"Priorização do Ciclo 1 → URC","acoes":["CAN-13","CAN-12"],"guardiao":"Comitê",
  "fallback":"Critérios como primeira pauta do Comitê após o início do ciclo — custa zero proteger. Sem critérios prévios, a triagem reabre discussão com demandas já na mesa.",
  "gatilho":"Critérios não aprovados até 21/08 → âmbar no painel e pauta obrigatória."},
]
slas = [
 {"id":"SLA-URC","titulo":"Cronograma de entrega das demandas (URC)","acoes":["CAN-17"],
  "acao":"Negociar na própria devolutiva de 04/08: URC devolve cronograma em até 10 d.u. após cada lote priorizado. Sem isso, CAN-17 segue sendo uma linha com dono 'URC' e nada mais."},
 {"id":"SLA-SOL","titulo":"Pedido das 4 oficinas (Soluções)","acoes":["PTF-01","PTF-02"],
  "acao":"O pedido de 14/08 sai completo: proposta de formato junto e datas propostas de 28/09 a 09/10, fora da colisão com o fim do Ciclo 2."},
]
folga = ("Janelas reais: 26/08–04/09 e pós-16/10. Se URC e UDT confirmarem datas até 14/08, o ano fecha com ~2–3 semanas de folga; "
         "se qualquer um deslizar para setembro, a régua de dezembro vira 'encaminhamentos claros' em vez de implantação iniciada.")

canais = [
 {"id":"cnr","nome":"CNR","completo":"CNR: operação ativa e receptiva","formato":"2 sessões · 3h · presencial",
  "pauta":["Projeto, UFs presentes e time","NPS e volume de ligações/dia; da dúvida tributária ao apoio na guia DAS","Valor mesmo para projetos sazonais, pequenos ou MVP (case UI × CNR na MEE)"]},
 {"id":"assessoria","nome":"Assessoria de Negócios","completo":"Assessoria de negócios","formato":"2 sessões · 3h · presencial",
  "pauta":["Projeto, UFs presentes e time","Assessoria como promotora das estratégias da UI","Casos de relacionamento convertido em consumo de solução; conexão com o Foco"]},
 {"id":"portal","nome":"Portal","completo":"Portal Sebrae","formato":"2 sessões · 3h · presencial",
  "pauta":["O que mudou do Portal 2025 para 2026","Integração Portal × MKT × Foco para leads (NA ou UFs)","Banner com QR code das páginas dos projetos em eventos","Área logada × não logada; Portal como centralizador","O que o Portal executa mas é gerido por Soluções"]},
 {"id":"loja","nome":"Loja","completo":"Loja Sebrae","formato":"2 sessões · 3h · presencial",
  "pauta":["Maturidade, adoção pelas UFs, receita gerada","Loja vs. Sympla e similares: comparativo com diferenciais (AMEI, Foco, Portal, captura em tempo real sem atrito)"]},
 {"id":"mkt","nome":"Marketing Cloud","completo":"CRM · Marketing Cloud","formato":"2 sessões · 3h · presencial",
  "pauta":["LPs, e-mail, SMS, WhatsApp; Bureau de CRM e jornadas","Disparo × campanha; transbordo entre canais e rastreabilidade","Campanhas do NA com alcance nacional","O que trava pedidos: sem estratégia, sem transbordo, sem base qualificada"]},
 {"id":"foco","nome":"Foco+","completo":"CRM · Foco+","formato":"2 sessões · 3h · presencial",
  "pauta":["Foco como centralizador dos dados e repositório dos sistemas externos","Além do atendimento: grupos de clientes, credenciamento, histórico, cruzamentos","Volume de informações e principais módulos"]},
 {"id":"rede","nome":"Rede própria e parceira","completo":"Rede própria e rede parceira","formato":"2 sessões · 3h · presencial",
  "pauta":["Mais de 14.000 pessoas onde não há escritório do Sebrae","Como a rede potencializa programas das UFs e do NA"]},
 {"id":"dxp","nome":"DXP","completo":"Estratégia DXP","formato":"formato reduzido · 2h · candidato a online",
  "pauta":["O que é a estratégia DXP","Como pode transformar o relacionamento da Inovação com clientes e parceiros"]},
]

ciclos = [
 {"id":"c1","nome":"Ciclo 1","janela":"11 a 21/08","ini":"2026-08-11","fim":"2026-08-21","regra":"Máximo de 1 apresentação por dia","obs":"1ª sessão de cada canal"},
 {"id":"intervalo","nome":"Intervalo de ajuste","janela":"24 a 28/08","ini":"2026-08-24","fim":"2026-08-28","regra":"URC e UI ajustam apresentações e dinâmicas com base na 1ª rodada","obs":""},
 {"id":"c2","nome":"Ciclo 2","janela":"31/08 a 11/09","ini":"2026-08-31","fim":"2026-09-11","regra":"Máximo de 1 apresentação por dia","obs":"2ª sessão de cada canal"},
]
encontros = []
for c in ("c1","c2"):
    for ca in canais:
        encontros.append({"id":f"{c}-{ca['id']}","ciclo":c,"canal":ca["id"],"data":None,"hora":None,
                          "local":None,"modo":"presencial" if ca["id"]!="dxp" else "a definir",
                          "status":"a agendar","confirmados":0,"convidados":0,"nota":""})

iniciativas = [{"nome":n,"nucleo":""} for n in [
 "ACT Ministério da Saúde","ALI Academy","ALI Coop","ALI Ecossistema","ALI IG","ALI Produtividade","ALI Rural",
 "ATIVA","Catalisa Gov","Catalisa ICT","Consult","Convênio Anprotec","ELI","ELI Summit","Embrapii",
 "Impulso Tecnológico","Inova Biomas","Internacionalização","MEI + Inovador","Missões","Salas do Empreendedor",
 "Sebrae Origens","Sebrae Startups","Sebraetec","Startup Day","Startup NE","Websummit Rio e Lisboa"]]
assert len(iniciativas) == 27

# matriz nasce limpa: estados '' | previsto | oficina | formulario | justificado | priorizado | encaminhado
matriz = {i["nome"]: {c["id"]: "" for c in canais} for i in iniciativas}

pessoas = [
 {"nome":"JR. (José Júnior)","papel":"Coordenação do plano · ponto de contato com a URC","grupo":"UI"},
 {"nome":"Gerência UI","papel":"Guardiã dos Nós 1 e 4 · comunicação institucional","grupo":"UI"},
 {"nome":"Sandra","papel":"Assistente do plano: agendas, prazos, monitoramento, boletim","grupo":"UI"},
 {"nome":"Anny","papel":"Comunicação · convite e articulação da oficina UDT","grupo":"UI"},
 {"nome":"Pova","papel":"Agente de IA de apoio ao formulário (CAN-08/09)","grupo":"UI"},
 {"nome":"Comitê de Atendimento e Relacionamento da Inovação","papel":"Coordenadores + representantes dos núcleos · curadoria e priorização","grupo":"Comitê"},
 {"nome":"Enio","papel":"URC · liderança (destinatário da proposta)","grupo":"URC"},
 {"nome":"Milva","papel":"URC · liderança (destinatária da proposta)","grupo":"URC"},
 {"nome":"Representante Núcleo 1","papel":"A indicar até 07/08 (CAN-05) — fallback: gestor da iniciativa","grupo":"Núcleos","pendente":True},
 {"nome":"Representante Núcleo 2","papel":"A indicar até 07/08 (CAN-05) — fallback: gestor da iniciativa","grupo":"Núcleos","pendente":True},
 {"nome":"Representante Núcleo 3","papel":"A indicar até 07/08 (CAN-05) — fallback: gestor da iniciativa","grupo":"Núcleos","pendente":True},
 {"nome":"Representante Núcleo 4","papel":"A indicar até 07/08 (CAN-05) — fallback: gestor da iniciativa","grupo":"Núcleos","pendente":True},
 {"nome":"Representante Núcleo 5","papel":"A indicar até 07/08 (CAN-05) — fallback: gestor da iniciativa","grupo":"Núcleos","pendente":True},
]

config = {"projeto":"Meta Inovação 2026","subtitulo":"Atendimento e Relacionamento · Unidade de Inovação · Sebrae Nacional",
          "versao":"0.1.0","atualizado_em":"2026-08-11","hoje_referencia":None}
changelog = [{"data":"2026-08-11","versao":"0.1.0","nota":"Nasce o ambiente: dashboard, plano (47 ações com soluções), caminho crítico (7 nós + 2 SLAs), agenda dos ciclos, matriz de demandas 27×8, participantes e modo edição."}]

def emit(nome, chave, obj):
    with io.open(os.path.join(BASE, "data", nome), "w", encoding="utf-8") as f:
        f.write("window.DB = window.DB || {};\n")
        f.write(f"window.DB.{chave} = ")
        f.write(json.dumps(obj, ensure_ascii=False, indent=1))
        f.write(";\n")

emit("config.js","config",config)
emit("plano.js","plano",plano)
emit("nos.js","nos",{"nos":nos,"slas":slas,"folga":folga})
emit("canais.js","canais",canais)
emit("agenda.js","agenda",{"ciclos":ciclos,"encontros":encontros})
emit("iniciativas.js","iniciativas",iniciativas)
emit("matriz.js","matriz",matriz)
emit("pessoas.js","pessoas",pessoas)
emit("changelog.js","changelog",changelog)
print("dados gerados:", len(plano), "ações ·", len(iniciativas), "iniciativas ·", len(canais), "canais ·", len(nos), "nós ·", len(encontros), "encontros")
